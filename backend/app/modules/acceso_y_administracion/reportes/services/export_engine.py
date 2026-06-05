"""Exportación de resultados QBE (Excel, PDF, CSV) en memoria."""
from __future__ import annotations

import csv
from io import BytesIO, StringIO
from typing import Any

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def qbe_result_to_excel_bytes(result: dict[str, Any]) -> BytesIO:
    meta = result.get('meta') or {}
    rows: list[dict[str, Any]] = list(result.get('data') or [])

    columns: list[str] = list(meta.get('columns') or [])
    if not columns and rows:
        columns = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = 'Reporte'

    for col_idx, header in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=str(header))

    for row_idx, row_dict in enumerate(rows, start=2):
        for col_idx, key in enumerate(columns, start=1):
            val = row_dict.get(key)
            if val is not None and not isinstance(val, (str, int, float, bool)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(columns) + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = min(max(12, len(columns[col_idx - 1]) + 2), 40)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _qbe_columns_and_rows(result: dict[str, Any]) -> tuple[list[str], list[dict[str, Any]]]:
    meta = result.get('meta') or {}
    rows: list[dict[str, Any]] = list(result.get('data') or [])
    columns: list[str] = list(meta.get('columns') or [])
    if not columns and rows:
        columns = list(rows[0].keys())
    return columns, rows


def qbe_result_to_csv_bytes(result: dict[str, Any]) -> BytesIO:
    columns, rows = _qbe_columns_and_rows(result)
    text_buf = StringIO()
    writer = csv.writer(text_buf)
    writer.writerow(columns)
    for row_dict in rows:
        writer.writerow([_cell_str(row_dict.get(key)) for key in columns])
    raw = text_buf.getvalue().encode('utf-8-sig')
    buffer = BytesIO(raw)
    buffer.seek(0)
    return buffer


def _cell_str(val: Any) -> str:
    if val is None:
        return ''
    if isinstance(val, (str, int, float, bool)):
        return str(val)
    return str(val)


def qbe_result_to_pdf_bytes(result: dict[str, Any]) -> BytesIO:
    columns, rows = _qbe_columns_and_rows(result)
    meta = result.get('meta') or {}
    model_name = str(meta.get('model') or 'Reporte')
    total = meta.get('total_records', len(rows))

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=24,
        rightMargin=24,
        topMargin=28,
        bottomMargin=28,
    )
    styles = getSampleStyleSheet()
    story: list[Any] = [
        Paragraph(f'<b>Reporte — {model_name}</b>', styles['Title']),
        Paragraph(f'Registros exportados: {len(rows)} (total consulta: {total})', styles['Normal']),
        Spacer(1, 12),
    ]

    if not columns:
        story.append(Paragraph('Sin columnas para mostrar.', styles['Normal']))
    else:
        table_data = [columns]
        for row_dict in rows:
            table_data.append([_cell_str(row_dict.get(key)) for key in columns])
        table = Table(table_data, repeatRows=1)
        table.setStyle(
            TableStyle(
                [
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
                ]
            )
        )
        story.append(table)

    doc.build(story)
    buffer.seek(0)
    return buffer
